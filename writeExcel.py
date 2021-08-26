import pandas
from openpyxl import load_workbook

# Function that opens the excel file and appends the data into the sheet
def writeExcel(filename, df1):
    book = load_workbook(filename)
    writer = pandas.ExcelWriter(filename, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        df1.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

    writer.save()